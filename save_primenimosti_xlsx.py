import os.path
import openpyxl
from main_window_voltage import *

def formirovanie_cross(cross):
    spisok_cross = []
    for katalog, nomer in cross.items():
        if not isinstance(nomer, str):  # тут если номер не один
            for nomer1 in nomer:
                spisok_cross.append([katalog, nomer1])
        else:
            spisok_cross.append([katalog, nomer])
    return spisok_cross

def formirovanie_cross_cross(spisok_cross):
    result_cross = []
    len_spisok_cross = len(spisok_cross)
    for i in range(len_spisok_cross):
        for n in range(i + 1, len_spisok_cross):
            # result_cross.append([set(spisok_cross[1]), set(spisok_cross[n])])
            result_cross.append([*spisok_cross[i], *spisok_cross[n]])
    return result_cross

# запись данных в файл
def save_dannix_detali(nomer_grout, cross, katalog='KRAUF'):

    new_cross = formirovanie_cross(cross)
    res = formirovanie_cross_cross(new_cross)
    soobsenie_v_formu = ''
    try:
        if os.path.isfile(f'voltage.xlsx'):  # Если файл сужествует открываем для записи
            print('Файл уже есть, открываем')
            soobsenie_v_formu += 'Файл voltage.xlsx уже есть, открываем. '
            excel_file = openpyxl.load_workbook(f'voltage.xlsx')
            shet_names = excel_file.sheetnames
            if nomer_grout in shet_names:  # проверияем существует ли лист с такой деталью
                print(f'Есть такой лист. Пересохраняем')
                soobsenie_v_formu += f'Лист с деталью {nomer_grout} уже есть, пересохраняем. '
                #excel_sheet = excel_file.create_sheet(title=(f'{list}NEW'))
                #excel_sheet = excel_file[model]
                excel_file.remove(excel_file[nomer_grout])
                excel_file.save('voltage.xlsx')
                excel_sheet = excel_file.create_sheet(title=nomer_grout)
            else:
                excel_sheet = excel_file.create_sheet(title=nomer_grout)
                print('Сохраняем новый лист')
                soobsenie_v_formu += f'Сохраняем деталь с номером {nomer_grout} на новый лист. '
        else:  # Иначе открываем пустой и формуем лист
            excel_file = openpyxl.Workbook()
            excel_sheet = excel_file.active
            excel_sheet.title = nomer_grout
            print('Новый файл, новый лист')
            soobsenie_v_formu += f'Файла voltage.xlsx не существует, открываем новый.  Сохраняем деталь с номером {nomer_grout} на новый лист.'

        # Установки ширины столбцов
        excel_sheet.column_dimensions["A"].width = 25
        excel_sheet.column_dimensions["B"].width = 25
        excel_sheet.column_dimensions["C"].width = 25
        excel_sheet.column_dimensions["D"].width = 25
        excel_sheet.column_dimensions["E"].width = 25
        # Запись сроссов в файл
        stroka = 1
        excel_sheet.cell(row=stroka, column=1).value = 'Каталог'
        excel_sheet.cell(row=stroka, column=2).value = 'Номер'
        excel_sheet.cell(row=stroka, column=3).value = 'Каталог1'
        excel_sheet.cell(row=stroka, column=4).value = 'Номер1'
        excel_sheet.cell(row=stroka, column=5).value = 'Источник'
        stroka += 1
        for stroka_crossov in res:
            katalog1, nomer1, katalog2, nomer2 = stroka_crossov# .split(',')
            excel_sheet.cell(row=stroka, column=1).value = katalog1
            excel_sheet.cell(row=stroka, column=2).value = nomer1
            excel_sheet.cell(row=stroka, column=3).value = katalog2
            excel_sheet.cell(row=stroka, column=4).value = nomer2
            excel_sheet.cell(row=stroka, column=5).value = f"сайт Вольтаж, номер-' {nomer_grout}"
            stroka += 1
        excel_file.save(f'voltage.xlsx')
    except Exception as error:
        print('Ошибка в формировании и сохранении файла: ' + repr(error))
        soobsenie_v_formu += f'Ошибка в формировании и сохранении файла: {repr(error)}'
    return soobsenie_v_formu


if __name__ == '__main__':
    cross = {'AS': 'S0422', 'Bosch': ['0001218141', '0001218741', '0986015930', '0986015931'], 'Cargo': ['111421', '116327'], 'Citroen': '5802000', 'Delco': 'DRS5930', 'FIAT': ['46231644', '7589005', '7589151', '75891510', '7726946', '77679430'], 'HC-PARTS': 'CS554', 'KRAUF': 'STB0554MN', 'LUCAS': 'LRS743', 'Magneti Marelli': ['063222034010', '63222034', '63222138', 'E95R22KW12V'], 'Motorherz': ['STB0554WA', 'STE0554RB'], 'Prestolite': '35260920', 'Valeo': '455626', 'WAI': '30743R', 'ZAUFER': '300N10910Z'}
    nomer_grout = 'STB0554'
    # save_dannix_detali(nomer_grout, cross, 'KRAUF')
    save_dannix_detali('SDE2818', cross, 'KRAUF')


